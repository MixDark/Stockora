import io
from decimal import Decimal

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from inventory.models import StockItem, StockMovement, Warehouse
from products.models import Product
from sales.models import Sale, Return


# ─── Helpers ────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill('solid', fgColor='4F46E5')   # indigo-600
ALT_ROW_FILL = PatternFill('solid', fgColor='EEF2FF')   # indigo-50
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
BOLD_FONT    = Font(bold=True, size=11)
THIN_BORDER  = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)


def _apply_table_style(ws, headers, rows, col_widths=None):
    """Escribe encabezados + filas con estilo de tabla en la hoja ws."""
    # Encabezados
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Datos
    for row_idx, row in enumerate(rows, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    # Anchos de columna
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Freeze primera fila
    ws.freeze_panes = 'A2'


def _excel_response(wb, filename):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Vistas ─────────────────────────────────────────────────────────────────

@login_required
def reports_home(request):
    return render(request, 'reports/home.html')


@login_required
def inventory_report(request):
    warehouse_id = request.GET.get('warehouse', '')
    items = StockItem.objects.select_related('product', 'warehouse', 'product__category')
    if warehouse_id:
        items = items.filter(warehouse_id=warehouse_id)
    items = list(items)
    total_value = sum(item.quantity * item.product.cost_price for item in items)
    total_products = len(items)

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario'

        headers = ['SKU', 'Producto', 'Categoría', 'Almacén', 'Stock', 'Costo unit.', 'Valor total', 'Estado']
        rows = []
        for item in items:
            if item.quantity == 0:
                estado = 'Sin stock'
            elif item.is_low_stock:
                estado = 'Stock bajo'
            else:
                estado = 'OK'
            rows.append([
                item.product.sku,
                item.product.name,
                item.product.category.name if item.product.category else '—',
                item.warehouse.name,
                item.quantity,
                float(item.product.cost_price),
                float(item.quantity * item.product.cost_price),
                estado,
            ])

        _apply_table_style(ws, headers, rows, col_widths=[14, 30, 18, 20, 10, 14, 16, 12])

        # Fila de total
        total_row = ws.max_row + 1
        total_cell = ws.cell(row=total_row, column=6, value='TOTAL')
        total_cell.font = BOLD_FONT
        value_cell = ws.cell(row=total_row, column=7, value=float(total_value))
        value_cell.font = BOLD_FONT

        return _excel_response(wb, f'inventario_{timezone.now().strftime("%Y%m%d")}.xlsx')

    if request.GET.get('export') == 'pdf':
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=1*cm, leftMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=14, spaceAfter=4)
        sub_style   = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#6B7280'), spaceAfter=10)
        elements.append(Paragraph('Reporte de Inventario', title_style))
        elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', sub_style))

        header = ['SKU', 'Producto', 'Categoría', 'Stock', 'Costo unit.', 'Valor total', 'Estado']
        data = [header]
        for item in items:
            if item.quantity == 0:
                estado = 'Sin stock'
            elif item.is_low_stock:
                estado = 'Stock bajo'
            else:
                estado = 'OK'
            data.append([
                item.product.sku,
                item.product.name[:35],
                item.product.category.name if item.product.category else '—',
                str(item.quantity),
                f'${item.product.cost_price:,.0f}',
                f'${item.quantity * item.product.cost_price:,.0f}',
                estado,
            ])
        data.append(['', '', '', '', 'TOTAL', f'${total_value:,.0f}', ''])

        col_widths = [2.5*cm, 7*cm, 4*cm, 2*cm, 3.5*cm, 4*cm, 3*cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN',      (1, 1), (2, -1), 'LEFT'),
            ('FONTSIZE',   (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EEF2FF')]),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#D1D5DB')),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EEF2FF')),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHT',  (0, 0), (-1, -1), 18),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="inventario_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response

    return render(request, 'reports/inventory.html', {
        'items': items, 'total_value': total_value, 'total_products': total_products,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'selected_warehouse': warehouse_id,
        'stock_items': items,
    })


@login_required
def sales_report(request):
    today = timezone.now().date()
    date_from = request.GET.get('date_from', (today - timedelta(days=30)).isoformat())
    date_to   = request.GET.get('date_to',   today.isoformat())

    sales = Sale.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        status=Sale.STATUS_COMPLETED,
    ).select_related('customer', 'warehouse')

    sales_list = list(sales)
    total_sales   = len(sales_list)
    total_revenue = sum(s.total for s in sales_list)
    avg_ticket    = (total_revenue / total_sales) if total_sales else Decimal('0')
    total_returns = Return.objects.filter(
        sale__created_at__date__gte=date_from, sale__created_at__date__lte=date_to
    ).count()

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ventas'

        headers = ['Factura', 'Fecha', 'Cliente', 'Almacén', 'Método de pago', 'Subtotal', 'IVA', 'Descuento', 'Total', 'Estado']
        rows = []
        for sale in sales_list:
            rows.append([
                sale.invoice_number,
                sale.created_at.strftime('%d/%m/%Y %H:%M'),
                sale.customer.name if sale.customer else 'Consumidor final',
                sale.warehouse.name,
                sale.get_payment_method_display(),
                float(sale.subtotal),
                float(sale.tax_total),
                float(sale.discount),
                float(sale.total),
                sale.get_status_display(),
            ])

        _apply_table_style(ws, headers, rows, col_widths=[22, 18, 25, 20, 18, 14, 10, 12, 14, 14])

        # Fila de totales
        r = ws.max_row + 1
        ws.cell(row=r, column=8, value='TOTAL').font = BOLD_FONT
        ws.cell(row=r, column=9, value=float(total_revenue)).font = BOLD_FONT

        # Hoja de KPIs
        ws2 = wb.create_sheet('Resumen')
        ws2.append(['Métrica', 'Valor'])
        ws2.cell(1, 1).font = HEADER_FONT; ws2.cell(1, 1).fill = HEADER_FILL
        ws2.cell(1, 2).font = HEADER_FONT; ws2.cell(1, 2).fill = HEADER_FILL
        ws2.append(['Ventas totales', total_sales])
        ws2.append(['Facturación total', float(total_revenue)])
        ws2.append(['Ticket promedio', float(avg_ticket)])
        ws2.append(['Devoluciones', total_returns])
        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 18

        return _excel_response(wb, f'ventas_{date_from}_{date_to}.xlsx')

    return render(request, 'reports/sales.html', {
        'sales': sales_list, 'date_from': date_from, 'date_to': date_to,
        'total_sales': total_sales, 'total_revenue': total_revenue,
        'avg_ticket': avg_ticket, 'total_returns': total_returns,
    })


@login_required
def movements_report(request):
    today = timezone.now().date()
    date_from     = request.GET.get('date_from', (today - timedelta(days=30)).isoformat())
    date_to       = request.GET.get('date_to',   today.isoformat())
    movement_type = request.GET.get('movement_type', '')

    movements = StockMovement.objects.filter(
        created_at__date__gte=date_from, created_at__date__lte=date_to,
    ).select_related('product', 'from_warehouse', 'to_warehouse', 'performed_by')
    if movement_type:
        movements = movements.filter(movement_type=movement_type)

    movements_list = list(movements)

    if request.GET.get('export') == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Movimientos'

        headers = ['Fecha', 'Tipo', 'Producto', 'SKU', 'Almacén origen', 'Almacén destino', 'Cantidad', 'Costo unit.', 'Referencia', 'Realizado por']
        rows = []
        for mv in movements_list:
            rows.append([
                mv.created_at.strftime('%d/%m/%Y %H:%M'),
                mv.get_movement_type_display(),
                mv.product.name,
                mv.product.sku,
                mv.from_warehouse.name if mv.from_warehouse else '—',
                mv.to_warehouse.name   if mv.to_warehouse   else '—',
                mv.quantity,
                float(mv.unit_cost),
                mv.reference or '—',
                mv.performed_by.get_full_name() if mv.performed_by else '—',
            ])

        _apply_table_style(ws, headers, rows, col_widths=[18, 14, 28, 14, 20, 20, 10, 12, 20, 22])
        return _excel_response(wb, f'movimientos_{date_from}_{date_to}.xlsx')

    return render(request, 'reports/movements.html', {
        'movements': movements_list, 'date_from': date_from, 'date_to': date_to,
        'selected_type': movement_type,
    })

